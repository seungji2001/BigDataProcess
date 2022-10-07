#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1;
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
	row_id += 1

#grade
student_cnt = ws.max_row - 1
grade_a = int(student_cnt * 0.3)
grade_a1 = int(student_cnt * 0.3 * 0.5)

grade_b = int(student_cnt * 0.7) 
grade_b1 = int((grade_b - grade_a) * 0.5) + grade_a

grade_c = student_cnt
grade_c1 = int((grade_c - grade_b) * 0.5) + grade_b

row_id = 1;
total = []
for row in ws:
	if row_id != 1:
		grade = ws.cell(row=row_id,column = 7).value
		total.append(grade)
	row_id += 1

total.sort()

row_id = 1;
for row in ws:	
	if row_id != 1:
		if ws.cell(row=row_id,column=7).value >= total[student_cnt - grade_a1]:
			ws.cell(row=row_id,column=8,value="A+")
		elif ws.cell(row=row_id,column = 7).value >= total[student_cnt - grade_a]:
			ws.cell(row=row_id,column=8,value="A0")
		elif ws.cell(row=row_id,column=7).value >= total[student_cnt - grade_b1]:
			ws.cell(row=row_id,column=8,value="B+")
		elif ws.cell(row=row_id,column=7).value >= total[student_cnt - grade_b]:
			ws.cell(row=row_id,column=8,value="B0")
		elif ws.cell(row=row_id,column=7).value >= total[student_cnt - grade_c1]:
			ws.cell(row=row_id,column=8,value="C+")
		else:
			ws.cell(row=row_id,column=8,value="C0")
	row_id+=1

			
wb.save("student.xlsx")
